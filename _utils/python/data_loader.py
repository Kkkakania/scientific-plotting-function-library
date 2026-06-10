"""数据载入助手：CSV / Excel / MAT / NumPy 文件统一接口.

把外部数据规整成模板期望的形状，省去每次手动 reshape。

用法::

    from data_loader import load_xy, load_matrix, load_groups

    x, y = load_xy('measure.csv', x_col='time', y_col='voltage')
    M = load_matrix('heatmap.xlsx', sheet='Sheet1')
    labels, values = load_groups('survey.csv', label_col='category', value_col='count')
"""
from pathlib import Path
import numpy as np


# ============ 单条 (x, y) ============

def load_xy(path, x_col=0, y_col=1, **opts):
    """读取一列 x、一列 y。

    支持: .csv / .tsv / .xlsx / .xls / .mat / .npy / .npz

    x_col / y_col 可以是列名（字符串）或列索引（整数）。
    """
    p = Path(path); suffix = p.suffix.lower()

    if suffix in {'.csv', '.tsv'}:
        sep = ',' if suffix == '.csv' else '\t'
        import csv
        rows = []
        with open(p, encoding=opts.get('encoding', 'utf-8')) as f:
            reader = csv.reader(f, delimiter=sep)
            for row in reader:
                rows.append(row)
        header = rows[0]
        data = rows[1:] if not isinstance(x_col, int) or not _all_numeric(rows[0]) else rows
        x_idx = header.index(x_col) if isinstance(x_col, str) else x_col
        y_idx = header.index(y_col) if isinstance(y_col, str) else y_col
        x = np.array([float(r[x_idx]) for r in data])
        y = np.array([float(r[y_idx]) for r in data])
        return x, y

    if suffix in {'.xlsx', '.xls'}:
        import openpyxl
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb[opts.get('sheet', wb.sheetnames[0])]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        x_idx = header.index(x_col) if isinstance(x_col, str) else x_col
        y_idx = header.index(y_col) if isinstance(y_col, str) else y_col
        x = np.array([r[x_idx] for r in rows[1:] if r[x_idx] is not None], dtype=float)
        y = np.array([r[y_idx] for r in rows[1:] if r[y_idx] is not None], dtype=float)
        return x, y

    if suffix == '.mat':
        from scipy.io import loadmat
        d = loadmat(p)
        return np.array(d[x_col]).ravel(), np.array(d[y_col]).ravel()

    if suffix == '.npy':
        a = np.load(p)
        return a[:, x_col] if isinstance(x_col, int) else None, \
               a[:, y_col] if isinstance(y_col, int) else None

    if suffix == '.npz':
        d = np.load(p)
        return np.array(d[x_col]).ravel(), np.array(d[y_col]).ravel()

    raise ValueError(f'unsupported file: {suffix}')


# ============ 矩阵（热力图/等高线） ============

def load_matrix(path, **opts):
    """读取一个二维数值矩阵."""
    p = Path(path); suffix = p.suffix.lower()

    if suffix in {'.csv', '.tsv'}:
        sep = ',' if suffix == '.csv' else '\t'
        skip = opts.get('skip_header', 0)
        return np.genfromtxt(p, delimiter=sep, skip_header=skip)

    if suffix in {'.xlsx', '.xls'}:
        import openpyxl
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb[opts.get('sheet', wb.sheetnames[0])]
        rows = list(ws.iter_rows(values_only=True))
        skip = opts.get('skip_header', 0)
        return np.array([list(r) for r in rows[skip:]], dtype=float)

    if suffix == '.mat':
        from scipy.io import loadmat
        d = loadmat(p)
        var = opts.get('var')
        if var is None:
            keys = [k for k in d if not k.startswith('__')]
            var = keys[0]
        return np.array(d[var])

    if suffix == '.npy':
        return np.load(p)

    raise ValueError(f'unsupported file: {suffix}')


# ============ 分组数据（柱状/箱线） ============

def load_groups(path, label_col, value_col, **opts):
    """读取 (类别标签, 数值) 对.

    返回 (labels: list[str], values: np.ndarray)
    """
    p = Path(path); suffix = p.suffix.lower()

    if suffix in {'.csv', '.tsv'}:
        sep = ',' if suffix == '.csv' else '\t'
        import csv
        with open(p, encoding=opts.get('encoding', 'utf-8')) as f:
            rows = list(csv.reader(f, delimiter=sep))
        header = rows[0]
        l_idx = header.index(label_col) if isinstance(label_col, str) else label_col
        v_idx = header.index(value_col) if isinstance(value_col, str) else value_col
        labels = [r[l_idx] for r in rows[1:]]
        values = np.array([float(r[v_idx]) for r in rows[1:]])
        return labels, values

    if suffix in {'.xlsx', '.xls'}:
        import openpyxl
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb[opts.get('sheet', wb.sheetnames[0])]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        l_idx = header.index(label_col) if isinstance(label_col, str) else label_col
        v_idx = header.index(value_col) if isinstance(value_col, str) else value_col
        labels = [r[l_idx] for r in rows[1:] if r[l_idx] is not None]
        values = np.array([r[v_idx] for r in rows[1:] if r[v_idx] is not None], dtype=float)
        return labels, values

    raise ValueError(f'unsupported file: {suffix}')


# ============ 时间序列 ============

def load_timeseries(path, time_col=0, value_cols=None, **opts):
    """读取一个时间列 + 一组或多组数值列.

    value_cols 为 None 时取除时间列外所有数值列。
    返回 (t, Y) 其中 Y 是 (n_series, n_points)。
    """
    p = Path(path); suffix = p.suffix.lower()
    if suffix not in {'.csv', '.tsv', '.xlsx', '.xls'}:
        raise ValueError(f'unsupported file: {suffix}')

    if suffix in {'.csv', '.tsv'}:
        sep = ',' if suffix == '.csv' else '\t'
        import csv
        with open(p, encoding=opts.get('encoding', 'utf-8')) as f:
            rows = list(csv.reader(f, delimiter=sep))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb[opts.get('sheet', wb.sheetnames[0])]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

    header = rows[0]
    t_idx = header.index(time_col) if isinstance(time_col, str) else time_col
    if value_cols is None:
        value_idx = [i for i in range(len(header)) if i != t_idx]
    else:
        value_idx = [header.index(c) if isinstance(c, str) else c for c in value_cols]

    t = np.array([float(r[t_idx]) for r in rows[1:]])
    Y = np.array([[float(r[i]) for r in rows[1:]] for i in value_idx])
    return t, Y


# ============ helpers ============

def _all_numeric(row):
    try:
        for v in row: float(v)
        return True
    except (ValueError, TypeError):
        return False
